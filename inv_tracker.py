import openpyxl as xl
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import datetime
from settings import INVOICES_FOLDER, MANAGEMENT, TEMPLATE_FOLDER, DATE_FORMAT2, DATE_FORMAT, OCC_LIST, EXCEL_DATE_FORMAT, FORMULA_VAT, FORMULA_TOTAL, DUMPS_FOLDER, FORMULA_NO_NIGHTS, CURRENCY_FORMAT, FORMULA_MONTHLY_CHARGE
from Occupant import Occupant
import string
import calendar
import os
import json

MONTHS = ["january", "february", "march", "april", "may", "june", "july", "august", "september", "october", "november", "december"]
INVOICE_SHEET = "Brent TA Invoice (2)"

font_table = NamedStyle(name="font_table")
font_table.font = Font(size=12)
thin = Side(border_style="thin", color="000000")
font_table.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def open_tracker():
    wb = xl.load_workbook(filename=MANAGEMENT)

    # Worksheet is on default page
    ws = wb.active

    # Using generator. Skipping the first two rows as they are junk
    rows = ws.rows
    next(rows)
    next(rows)

    return rows

def full_populate() -> Occupant:

    rows = open_tracker()

    occupants : Occupant = []

    # x are the account holders
    # address, room no, name, ref, ~contact number, room size, start, end , pp night
    for x in rows:
        address = x[0].value
        if address is None:
            break

        occupants.append(Occupant(
            x[0], # address
            x[1], # room no
            x[2], # name
            x[3], # ref
            x[5], # room size
            x[6], # start date
            x[7], # end date
            x[8], # price per night
        ))

        # print(str(i+3) , x[2].value, x[6].value, x[7].value)


    return occupants

def save_log(dump : {}, name : str, loc : str):
    naming_log = str(datetime.datetime.now().strftime('log_'+name+'_%H_%M_%d_%m_%Y.json'))

    with open(DUMPS_FOLDER+loc+ naming_log, "w", encoding="utf-8") as f:
        json.dump(dump, f, ensure_ascii=False, indent=1)

def dumping_log(dump : {}, dump_iter : str, x : Occupant):
    dump[dump_iter] = [
        x.address.value,
        x.room.value,
        x.name.value,
        x.ref.value,
        x.room_size.value,
        str(x.start_date.value),
        str(x.end_date.value),
        x.rate.value,
        x.number_of_nights,
        str(x.cleaned_end),
    ]

def generate_occupancy_lists(arr : Occupant, debug=False) -> [list, list]:
    not_end = []
    end = []
    debug_ending = {}
    debug_not = {}

    for i, x in enumerate(arr):
        if x.end_occupancy():
            end.append(x)
            if debug:
                dumping_log(debug_ending, "ENDING: " + x.name.value + ": " + x.address.value + " -> ROW_INDEX: " + str(i+3), x)
        else:
            not_end.append(x)
            if debug:
                dumping_log(debug_not, "(NOT) ENDING: " + x.name.value + ": " + x.address.value + " -> ROW_INDEX: " + str(i+3), x)

    # print(debug_s)
    if debug:
        save_log(debug_ending, "OCCUPANCY_LIST", OCC_LIST+"ending/")
        save_log(debug_not, "(NOT)_OCCUPANCY_LIST", OCC_LIST+"not_ending/")

    return not_end, end

def open_invoice(name):
    wb = xl.load_workbook(filename=INVOICES_FOLDER+name)
    # ws = wb[INVOICE_SHEET]
    wb.add_named_style(font_table)
    # return ws
    return wb

# address, room no, room size, occupant, placement, start, end, no of nights, nightly rate ...
def compare_row_occupant(occupant : Occupant, row) -> bool:
    invoice = create_delete_invoice_object(row)
    if occupant.correct_invoice(invoice):
        return True
        # print(occupant.name.value, occupant.end_date.value, occupant.cleaned_end.month)
    return False
            
def delete_rows(sheet, idx: int, amount: int = 1):
    sheet.delete_rows(idx, amount)
    merged_cells = [_ for _ in sheet.merged_cells.ranges]
    for index, mcr in enumerate(merged_cells):
        if idx < mcr.min_row:
            if idx + amount - 1 >= mcr.min_row:
                mcr.shrink(top=idx + amount - mcr.min_row)
                if mcr.min_row > mcr.max_row:
                    sheet.merged_cells.ranges.remove(mcr)
                    continue
            mcr.shift(row_shift=-amount)
        elif idx <= mcr.max_row:
            mcr.shrink(bottom=min(mcr.max_row - idx + 1, amount))
        if mcr.min_row > mcr.max_row:
            sheet.merged_cells.ranges.remove(mcr)

def create_delete_invoice_object(row) -> Occupant:
    occ = Occupant(
        row[0],  # address
        row[2],  # room no
        row[4],  # occupant
        row[5],  # ref
        row[3],  # room size
        # row[5].value, # start
        "00/00/00",
        row[7],  # end
        row[9]  # rate
        # row[7].value  # nights
    )

    occ.end_occupancy()
    return occ

def clean_with_end_date(occupants, workbook):
    ws = workbook[INVOICE_SHEET]
    worksheet_month = retrieve_invoice_month(ws)

    for occupant in occupants:
        if occupant.end_occupancy():
            for x in ws["E"]:
                # clean DATA with lstrip()
                if x.value == occupant.name.value.lstrip().rstrip():
                    if compare_row_occupant(occupant, ws[x.row]):
                        if occupant.need_to_delete_invoice(worksheet_month):
                            # print("DELETE ROW: ", occupant.name.value)
                            delete_rows(ws, x.row, 1)
                        else:
                            # print("UPDATE ROW: ", occupant.name.value, ws["H"+str(x.row)].value.strftime(DATE_FORMAT), "--> ", occupant.cleaned_end.strftime(DATE_FORMAT))
                            ws["H" + str(x.row)].value = occupant.cleaned_end
                            
    fix_formulas(ws)

    workbook.save("invoices/september22Outcome.xlsx")

def fix_formulas(ws):

    # Number of nights
    for nights in ws["I"]:
        val = nights.value
        if val is None:
            continue
        if nights.value[0] == "=":
            row_num = str(nights.row)
            nights.value = FORMULA_NO_NIGHTS.format(row_num, row_num)

    update_formula_tallys(ws, "K", FORMULA_MONTHLY_CHARGE)
    update_formula_tallys(ws, "L", FORMULA_VAT)
    update_formula_tallys(ws, "M", FORMULA_TOTAL)

def update_formula_tallys(ws_col, letter : str, formula : str):
    for cell in ws_col[letter]:
        val = cell.value
        if val is None:
            continue
        if val[:4] == "=SUM":
            cell.value = "=SUM("+letter+"1"+":"+letter+str(cell.row-1)+")"
            continue
        if val[0] == "=":
            row_num = str(cell.row)
            cell.value = formula.format(row_num, row_num)

def retrieve_invoice_month(worksheet) -> int:
    month = str(worksheet["B6"].value).lower()
    for i, mon in enumerate(MONTHS):
       if mon == month:
           return i+1

    return 1

def replace_date_col(ws_col, year : int, month : int, date : int):
    store = datetime.datetime(year, month, date)
    for cell in ws_col:
        if type(cell).__name__ == "MergedCell":
            continue
        if cell.is_date and cell.value is not None:
            cell.value = store
    return store

def fix_date_cells(ws, f_cell, s_cell):
    temp = s_cell.value
    ws.unmerge_cells(f_cell.coordinate+":"+s_cell.coordinate)
    ws[s_cell.coordinate].number_format = EXCEL_DATE_FORMAT
    ws[s_cell.coordinate].font = ws[s_cell.coordinate].font.copy(size=12)
    ws[s_cell.coordinate].value = temp

def fix_merged_cells_dates(ws, to_merge, first_cell):
    for cell in ws[to_merge]:
        # print(ws[first_cell+str(cell.row)].value)
        f_cell = ws[first_cell+str(cell.row)]
        # if f_cell.value == "Rental Period".lower():
        if str(f_cell.value).lower() == "Rental Period":
            # print("Need to merge here")
            ws.merge_cells(first_cell+str(cell.row)+":"+to_merge+str(cell.row))
        elif f_cell.is_date:
            for mc in ws.merged_cells.ranges:
                if f_cell.coordinate in mc:
                    fix_date_cells(ws, f_cell, ws[to_merge+str(cell.row)])

def fix_merged_cells_address(ws, merge_col, first_col):
    for cell in ws[merge_col]:
        if type(cell).__name__ == "MergedCell":
            ws.merge_cells(first_col+str(cell.row)+":"+cell.coordinate)

# month -> invoice month    year -> invoice year   date -> tracker occupancy start date
def determine_invoice_start_date(month, year, date_string):
    date_string = date_string.lstrip().rstrip()
    date : datetime

    try:
        date = datetime.datetime.strptime(date_string, DATE_FORMAT)
    except:
        try:
            date = datetime.datetime.strptime(date_string, DATE_FORMAT2)
        except:
            return datetime.datetime(year, month, 1)

    if date.month < month or date.year < year:
        return datetime.datetime(year, month, 1)
    return date

def num_days_month(month) -> int:
    return calendar.monthrange(2022, month)[1]

def update_font_style(ws, row_num, letters):
    row = str(row_num)
    # for letter in letters:
    #     ws[letter + row].font = ws[letter + row].font.copy(size=12)
    for cell in ws[row+":"+row]:
        if get_column_letter(cell.column) == "N":
            break
        cell.style = "font_table"

def check_end_append_conditions(month, year, date):
    if date.month == month and date.year == year:
        return True
    return False

# worksheet, row number to insert info, data occupant, address, last date of month
def insert_occupant_row_information(ws, row_num, occupant, address, last_day_month):

    update_font_style(ws, row_num, list(string.ascii_uppercase))

    # address
    ws["A"+str(row_num)].value = str(address).lstrip().rstrip()
    ws.merge_cells("A"+str(row_num)+":"+"B"+str(row_num))
    # ws["A" + str(row_num)].font = ws["A" + str(row_num)].font.copy(size=12)
    # room no
    ws["C"+str(row_num)].value  = occupant.room.value
    # room size
    ws["D"+str(row_num)].value  = occupant.room_size.value
    # occupant
    ws["E"+str(row_num)].value  = str(occupant.name.value).lstrip().rstrip()
    # ref
    ws["F"+str(row_num)].value  = occupant.ref.value
    # start
    # CHECK START DATE. NEEDS TO BE PARSED AS A DATE
    # print(occupant.start_date.value)
    ws["G"+str(row_num)].value = determine_invoice_start_date(last_day_month.month, last_day_month.year, occupant.start_date.value)
    # ws["G" + str(row_num)].value = datetime.datetime(2022, 9, 1)
    ws["G" + str(row_num)].number_format = EXCEL_DATE_FORMAT
    # end
    ws["H"+str(row_num)].value  = last_day_month
    ws["H"+str(row_num)].number_format = EXCEL_DATE_FORMAT
    # fix_date_cells(ws, f_cell, ws[to_merge + str(cell.row)])
    # no of nights
    ws["I"+str(row_num)].value  = FORMULA_NO_NIGHTS.format(row_num, row_num)
    # nightly rate
    try:
        ws["J"+str(row_num)].value  = int(float(occupant.rate.value[1:]))
        ws["J" + str(row_num)].number_format = CURRENCY_FORMAT
        # ws["J" + str(row_num)].alignment = Alignment(horizontal="center")
    except:
        # need to add to log
        pass
    # monthly charge
    ws["K"+str(row_num)].value  = FORMULA_MONTHLY_CHARGE.format(row_num, row_num)
    ws["K" + str(row_num)].number_format = CURRENCY_FORMAT
    # vat
    ws["L"+str(row_num)].value  = FORMULA_VAT.format(row_num)
    ws["L" + str(row_num)].number_format = CURRENCY_FORMAT
    # total
    ws["M"+str(row_num)].value  = FORMULA_TOTAL.format(row_num, row_num)
    ws["M" + str(row_num)].number_format = CURRENCY_FORMAT

def maintain_current_new(workbook, debug=False):
    ws = workbook[INVOICE_SHEET]

    occupants = full_populate()
    ws_month = retrieve_invoice_month(ws)

    # works out the date for the end of the month
    end_day = num_days_month(ws_month)

    # document must replace start and end date to the first/last day of the month respectively

    # occupants who are staying within this month, no further code is needed to update
    # replacing beginning date and end date
    replace_date_col(ws["G"], 2022, ws_month, 1)
    last_day_month = replace_date_col(ws["H"], 2022, ws_month, end_day)

    not_ending, ending = generate_occupancy_lists(occupants, debug=debug)

    # dictionary for saving information
    dump = {}
    dump_added = {}
    dump_iter = 0

    # we check those without an end date if they exist in the invoice
    # those who DO NOT, a new row must be appended with Occupancy start date and end of the month
    for x in not_ending:
        # if not x.end_occupancy():
        exists = False
        # checking for names
        for cell in ws["E"]:
            # if cell is None:
            #     continue
            # if str(cell.value).lstrip().rstrip() == x.name.value.lstrip().rstrip():
            if str(cell.value).lstrip().rstrip().lower() == x.name.value.lstrip().rstrip().lower():
                # we check if its the same, we can break from the loop
                if compare_row_occupant(x, ws[cell.row]):
                    exists = True
                    # print("EXISTS: " + x.name.value)
                    break
                else:
                    pass
                    # print("*********** ROW DOESNT EXISTS: " + x.name.value)
        # if it doesn't exist, we need to add row
        if not exists:
            # we need to find if that person already exists with same address so we can add another row around them
            # we also want to append a new row underneath the older entries

            addr_row = 0
            name_exists = False
            address_found = False

            # searching through address
            for cell in ws["A"]:
                # if address is here
                if x.compare_address(str(cell.value).lower()):
                # if x.compare_address(cell.value):
                    address_found = True
                    addr_row = cell.row

                    if x.compare_name(ws["E" + str(cell.row)].value.lower()):
                    # if x.compare_name(ws["E"+str(cell.row)].value):
                        name_exists = True
                        # print("----------Same NAME + ADDR")
                    elif addr_row > 0 and name_exists:
                        # print("*************APPEND UNDERNATH SAME NAME+ADDRESS: " + x.name.value)
                        ws.insert_rows(addr_row)
                        insert_occupant_row_information(ws, addr_row, x, cell.value, last_day_month)
                        if debug:
                            dumping_log(dump_added, x.name.value + ": " + x.address.value, x)
                        addr_row = 0
                        name_exists = False
                        address_found = True
                        break

                elif addr_row > 0:
                    # print("++++++++++++++APPEND AT THE BOTTOM: " + x.name.value)
                    ws.insert_rows(addr_row + 1)
                    insert_occupant_row_information(ws, addr_row + 1, x, ws[addr_row][0].value, last_day_month)
                    if debug:
                        dumping_log(dump_added, x.name.value +": " + x.address.value, x)
                    addr_row = 0
                    name_exists = False
                    address_found = True
                    break

            if not address_found and debug:
                # print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@ ADDRESS NOT FOUND: " + x.address.value)
                dump_iter += 1
                dumping_log(dump, str(dump_iter) + ": "+ x.address.value + ": " + x.name.value, x)

            address_found = False
            # if there is NO new name and or address for the occupant, then we append at the bottom of the list

            # if the address we are trying to find doesn't exist, then...
            # print("----------> UNIQUE ROW NOT EXISTS: " + x.name.value)


    # those who already exist, we do nothing

    # ----------------  CLEANING PART --------------------- #
    # NEED TO CHECK
    worksheet_month = retrieve_invoice_month(ws)

    dump_change_clean = {}
    dump_remove_clean = {}
    dump_clean_fail = {}

    for occupant in ending:
        found_at_least_once = False
        not_found = True
        # check if they are currently in the invoice
        for x in ws["E"]:
            if str(x.value).lower() == str(occupant.name.value).lstrip().rstrip().lower():
            # if x.value == occupant.name.value.lstrip().rstrip():
                found_at_least_once = True
                # delete row if in previous month
                if compare_row_occupant(occupant, ws[x.row]):
                    if occupant.need_to_delete_invoice(worksheet_month):
                        print("DELETE ROW: ", occupant.name.value)
                        delete_rows(ws, x.row, 1)
                        not_found = False
                        if debug:
                            naming_log = datetime.datetime.now().strftime('log_DEL_ROW_%H_%M_%d_%m_%Y.json')
                            dumping_log(dump_remove_clean, occupant.name.value + ": " + occupant.address.value, occupant)
                    # update ending date if they already exist
                    else:
                        print("UPDATE ROW: ", occupant.name.value, ws["H"+str(x.row)].value.strftime(DATE_FORMAT), "--> ", occupant.cleaned_end.strftime(DATE_FORMAT))
                        ws["H" + str(x.row)].value = occupant.cleaned_end
                        not_found = False
                        if debug:
                            dumping_log(dump_change_clean, occupant.name.value + ": " + occupant.address.value, occupant)

        # if they need to be added as a new entry and ending this month
        # if not_found and occupant.same_month(worksheet_month):
        if not_found and check_end_append_conditions(worksheet_month, 2022, occupant.cleaned_end):
            # print(occupant.name.value + ": Im here -> Month: " + str(occupant.cleaned_end.month) + " CE: " + str(occupant.cleaned_end))
            # print(occupant.name.value + ": Im here -> Month: " + str(occupant.end_date.value.month))
            for cell in ws["A"]:
                # if address is here
                if occupant.compare_address(cell.value):
                    # we will insert row above the first address
                    ws.insert_rows(cell.row)
                    insert_occupant_row_information(ws, cell.row-1, occupant, cell.value, occupant.cleaned_end)
                    break

        if not found_at_least_once and debug:
            dumping_log(dump_clean_fail, occupant.name.value + ": " + occupant.address.value, occupant)

    save_log(dump_added, "ADDED", "added/")
    save_log(dump_added, "ADDED", "added/")
    save_log(dump, "NOTFOUND", "not_found/")

    save_log(dump_remove_clean, "DEL_ROW", "maindel/del/")
    save_log(dump_change_clean, "UPDATING_END_DATE", "maindel/maintain/")
    save_log(dump_clean_fail, "NOT_FOUND_MAINDEL", "maindel/not_found/")


    fix_formulas(ws)

    fix_merged_cells_dates(ws, "H", "G")

    fix_merged_cells_address(ws, "B", "A")

    workbook.save("invoices/september22Outcome.xlsx")